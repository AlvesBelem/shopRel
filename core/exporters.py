from PyQt5.QtWidgets import QMessageBox
from openpyxl.styles import Font
from fpdf import FPDF
import pandas as pd


def exportar_excel(main_window, file_path):
    try:
        if not isinstance(main_window.df, pd.DataFrame) or main_window.df.empty:
            QMessageBox.warning(main_window, "Aviso", "Nenhum dado para exportar.")
            return

        total = main_window.df["cdcarregamento"].nunique()
        periodo = f"Período: {main_window.date_start.text()} a {main_window.date_end.text()}"

        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            main_window.df.to_excel(writer, index=False, sheet_name="Relatório", startrow=4)
            ws = writer.sheets["Relatório"]

            ws["A1"] = "Relatório de Entregas"
            ws["A2"] = periodo
            ws["A3"] = f"Total de carregamentos: {total}"

            ws["A1"].font = Font(size=14, bold=True)
            ws["A2"].font = Font(bold=True)
            ws["A3"].font = Font(bold=True)

            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max_length + 2

    except Exception as e:
        QMessageBox.critical(main_window, "Erro", f"Erro ao exportar Excel:\n{e}")


def exportar_pdf(main_window, file_path):
    try:
        if not isinstance(main_window.df, pd.DataFrame) or main_window.df.empty:
            QMessageBox.warning(main_window, "Aviso", "Nenhum dado para exportar.")
            return

        pdf = FPDF(orientation='L', unit='mm', format='A4')
        pdf.add_page()
        pdf.set_font("Helvetica", size=8)
        pdf.cell(0, 8, "Relatório de Entregas", ln=1)
        pdf.cell(0, 8, f"Período: {main_window.date_start.text()} a {main_window.date_end.text()}", ln=1, align='R')

        col_names = ["Carregamento", "Entregador", "Data Saída", "Data Retorno", "Status", "Clientes", "Documentos"]
        col_widths = [25, 45, 35, 35, 25, 55, 60]

        for i, col in enumerate(col_names):
            pdf.set_fill_color(200, 200, 200)
            pdf.cell(col_widths[i], 8, col, border=1, align='C', fill=True)
        pdf.ln()

        for _, row in main_window.df.iterrows():
            values = [
                row.get("cdcarregamento", ""),
                row.get("nmentregador", ""),
                row.get("data_saida", ""),
                row.get("data_retorno", ""),
                row.get("status_entrega", ""),
                row.get("nmpessoa", ""),
                row.get("documentos", "")
            ]
            for i, val in enumerate(values):
                text = str(val)[:col_widths[i] // 2] + ("..." if len(str(val)) > col_widths[i] // 2 else "")
                pdf.cell(col_widths[i], 8, text, border=1, align='C')
            pdf.ln()

        total = main_window.df["cdcarregamento"].nunique()
        pdf.ln(5) # type: ignore
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(0, 10, f"Total de carregamentos: {total}", ln=1, align="R")

        pdf.output(file_path)

    except Exception as e:
        QMessageBox.critical(main_window, "Erro", f"Erro ao gerar PDF:\n{e}")
